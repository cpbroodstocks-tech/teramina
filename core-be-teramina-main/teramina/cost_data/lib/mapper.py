# pylint: disable=C0209,R0914,R0915
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

def accounting_format(value):
    """to set the accounting format"""
    return "({:,.2f})".format(-value) if value < 0 else "{:,.2f}".format(value)

def generate_pl_report(farm_name, start_date, end_date, df):
    """pl report generator"""

    work_book = Workbook()
    work_sheet = work_book.active

    # Change width of column B
    work_sheet.column_dimensions["A"].width = 50
    work_sheet.column_dimensions["B"].width = 20

    # cells to merge
    work_sheet.merge_cells('A1:B1')
    work_sheet.merge_cells('A2:B2')
    work_sheet.merge_cells('A3:B3')

    company_cell = work_sheet.cell(row=1, column=1)
    company_cell.value = farm_name
    company_cell.alignment = Alignment(horizontal='center', vertical='center')
    company_cell.font = Font(bold=True, size=12, name="Arial")

    title_cell = work_sheet.cell(row=2, column=1)
    title_cell.value = 'Laba/Rugi'
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(bold=True, size=12, name="Arial")

    sub_title_cell = work_sheet.cell(row=3, column=1)
    sub_title_cell.value = f'{start_date} - {end_date}'
    sub_title_cell.alignment = Alignment(horizontal='center', vertical='center')
    sub_title_cell.font = Font(bold=False, size=11, name="Arial")

    categories = df["category"].unique()
    row = 4
    for cat in categories:
        row += 2
        category_df = df[df["category"] == cat]

        # print(cat, " : ", ndf["value"].sum())
        value_cell = work_sheet.cell(row=row, column=1)
        value_cell.value = cat
        value_cell.font = Font(bold=True, size=12, name="Arial")


        selected_parent_accounts = category_df["parent_account"].unique()
        for parent in selected_parent_accounts:
            row += 1
            parent_df = category_df[category_df["parent_account"] == parent]
            value_cell = work_sheet.cell(row=row, column=1)
            value_cell.value = parent
            value_cell.font = Font(bold=True, size=12, name="Arial")

            value_cell = work_sheet.cell(row=row, column=2)
            value_cell.value = accounting_format(parent_df["value"].sum())

            for item in parent_df["account"]:
                row += 1
                value_cell = work_sheet.cell(row=row, column=1)
                value_cell.value = item
                value_cell = work_sheet.cell(row=row, column=2)
                value_cell.value = accounting_format(
                    parent_df[parent_df["account"] == item]["value"].iloc[0]
                )

    pendapatan = df[df["category"] == "Pendapatan"]["value"].sum()
    hpp = df[df["category"] == "Harga Pokok Penjualan"]["value"].sum()
    biaya_operasi = df[df["category"] == "Beban Operasi"]["value"].sum()
    pendapatan_lain = df[df["parent_account"] == "PENDAPATAN DI LUAR USAHA"]["value"].sum()
    biaya_lain = df[df["parent_account"] == "BIAYA DI LUAR USAHA"]["value"].sum()
    total = pendapatan - (hpp + biaya_operasi) + (pendapatan_lain - biaya_lain)

    row += 2
    value_cell = work_sheet.cell(row=row, column=1)
    value_cell.value = "Laba-Rugi Bersih Sebelum Pajak"
    value_cell.font = Font(bold=True, size=12, name="Arial")
    value_cell = work_sheet.cell(row=row, column=2)
    value_cell.value = accounting_format(total)

    row += 1
    value_cell = work_sheet.cell(row=row, column=1)
    value_cell.value = "Laba-Rugi Bersih Setelah Pajak"
    value_cell.font = Font(bold=True, size=12, name="Arial")
    value_cell = work_sheet.cell(row=row, column=2)
    value_cell.value = accounting_format(111/100*total)

    return work_book
