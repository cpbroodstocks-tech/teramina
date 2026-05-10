"""Build a P&L Excel workbook using openpyxl."""
import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


_GRAY = PatternFill("solid", fgColor="F0F0F0")
_GREEN = PatternFill("solid", fgColor="E8F5E9")
_RED = PatternFill("solid", fgColor="FCE4EC")
_HEADER_FONT = Font(bold=True, size=12)
_SECTION_FONT = Font(bold=True, size=10)
_BOLD = Font(bold=True, size=10)
_NORMAL = Font(size=10)


def _idr(n):
    return int(n or 0)


def build_pl_excel(report: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L Report"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22

    row = [1]

    def write(label, value=None, bold=False, section=False, fill=None, align_right=True):
        r = row[0]
        ca = ws.cell(r, 1, label)
        cb = ws.cell(r, 2, value)
        if section:
            ca.font = _SECTION_FONT
            ca.fill = _GRAY
            cb.fill = _GRAY
        elif bold:
            ca.font = _BOLD
            cb.font = _BOLD
        else:
            ca.font = _NORMAL
            cb.font = _NORMAL
        if fill:
            ca.fill = fill
            cb.fill = fill
        if align_right and value is not None:
            cb.alignment = Alignment(horizontal="right")
        if isinstance(value, int):
            cb.number_format = '#,##0'
        row[0] += 1

    def gap():
        row[0] += 1

    # Title
    t = ws.cell(row[0], 1, "PROFIT & LOSS STATEMENT")
    t.font = _HEADER_FONT
    row[0] += 1
    ws.cell(row[0], 1, f"{report['cycle_name']} | {report['pond_name']} | {report['doc_range']}")
    row[0] += 1
    ws.cell(row[0], 1, f"Start: {report.get('start_date') or '-'}  |  Currency: {report['currency']}")
    row[0] += 1
    ws.cell(row[0], 1, f"Generated: {report['generated_at'][:19]}")
    row[0] += 2

    # Revenue
    write("REVENUE", section=True)
    for evt in report["harvest_events"]:
        h = "Final" if evt["harvest_type"] == "final" else f"Partial {evt['harvest_no']}"
        write(f"  {h} Harvest - DOC {evt['doc']} ({evt['biomass_kg']} kg)", _idr(evt["revenue_idr"]))
    if report.get("projected_remaining_idr") is not None:
        write("  Remaining in pond (projected)", _idr(report["projected_remaining_idr"]))
    write("Total Revenue", _idr(report["total_revenue_idr"]), bold=True)
    gap()

    # COGS
    write("COST OF PRODUCTION (COGS)", section=True)
    if report["cost_seed_idr"]:
        write("  Seed / Fry", _idr(report["cost_seed_idr"]))
    write("  Feed", _idr(report["cost_feed_idr"]))
    write("  Harvest Operations", _idr(report["cost_harvest_idr"]))
    write("Total COGS", _idr(report["total_cogs_idr"]), bold=True)
    gp_fill = _GREEN if (report["gross_profit_idr"] or 0) >= 0 else _RED
    write(f"Gross Profit ({report['gross_margin_pct']}%)", _idr(report["gross_profit_idr"]), bold=True, fill=gp_fill)
    gap()

    # OpEx
    write("OPERATING EXPENSES", section=True)
    write("  Labor", _idr(report["cost_labor_idr"]))
    write("  Energy", _idr(report["cost_energy_idr"]))
    write("  Probiotics & Treatment", _idr(report["cost_probiotics_idr"]))
    write("  Bonus", _idr(report["cost_bonus_idr"]))
    write("  Other", _idr(report["cost_other_idr"]))
    write("Total Operating Expenses", _idr(report["total_opex_idr"]), bold=True)
    gap()

    # Net
    write("NET RESULT", section=True)
    write("Total Cost", _idr(report["total_cost_idr"]), bold=True)
    np_fill = _GREEN if (report["net_profit_idr"] or 0) >= 0 else _RED
    write(f"Net Profit ({report['net_margin_pct']}%)", _idr(report["net_profit_idr"]), bold=True, fill=np_fill)
    gap()

    # KPIs
    write("KEY PERFORMANCE INDICATORS", section=True)
    kpi = report["kpi"]
    ws.column_dimensions["B"].width = 22
    write("  Cycle Duration (days)", kpi["doc"])
    write("  Total Harvest (kg)", kpi["total_harvest_kg"])
    write("  Final ABW (g)", kpi["final_abw_g"])
    write("  Survival Rate (%)", kpi["survival_rate_pct"])
    write("  FCR", kpi["fcr"])
    write("  Cost / kg (IDR)", _idr(kpi["cost_per_kg_idr"]))
    write("  Revenue / kg (IDR)", _idr(kpi["revenue_per_kg_idr"]))
    write("  Break-even Price (IDR/kg)", _idr(kpi["break_even_price_idr"]))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
